import openpyxl
from openpyxl.utils import get_column_letter


CELL_MAX_WIDTH = 300
CELL_PADDING = 5


def read_workbook(file_path):
    """ Returns the workbook instance for the specified file. """
    return openpyxl.load_workbook(file_path)


def read_workbook_columns(wb, sheet_index=0):
    """ Returns a list of columns for the specified workbook sheet. """
    sheet = wb._sheets[sheet_index]
    num_columns = sheet.max_column
    columns = []
    
    for i in range(1, num_columns + 1):
        cell_obj = sheet.cell(row=1, column=i)
        value = cell_obj.value
        if isinstance(value, str):
            value = value.strip()

        columns.append(value)

    return columns


def read_workbook_data(wb, sheet_index=0):
    """
    Returns the specified workbook sheet's data as an array of objects
    using the header columns as the keys
    """
    header_columns = read_workbook_columns(wb, sheet_index)
    sheet = wb._sheets[sheet_index]
    num_columns = sheet.max_column
    num_rows = sheet.max_row
    data = []
    
    for row in range(2, num_rows + 1):
        obj = {}
        for col in range(1, num_columns + 1):
            cell_obj = sheet.cell(row=row, column=col)
            value = cell_obj.value
            if isinstance(value, str):
                value = value.strip()
            elif value is None:
                value = ''
            obj[header_columns[col - 1]] = value
        
        data.append(obj)
    
    return data


def write_workbook_data(filename, sheets, data):
    """ Writes the specified data into the workbook. """
    wb = openpyxl.Workbook()
    
    for sheet_index, sheet_name in enumerate(sheets):
        wb.create_sheet(sheet_name, sheet_index)
        ws = wb.active
        col_widths = []
        
        for i, row in enumerate(data):
            for x, value in enumerate(row):
                cell = ws.cell(row=i + 1, column=x + 1)
                cell.value = value
                if isinstance(value, str):
                    new_width = len(cell.value) + CELL_PADDING
                    if new_width > CELL_MAX_WIDTH:
                        new_width = CELL_MAX_WIDTH

                    if len(col_widths) > x:
                        if new_width > col_widths[x]:
                            col_widths[x] = new_width
                    else:
                        col_widths.append(new_width)

        for i, col_width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = col_width
            
    wb.save(filename)
